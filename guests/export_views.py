"""
Export views for Guest Tracker
Handles CSV, Excel, and PDF exports
"""
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image as ReportLabImage
from reportlab.pdfgen import canvas
from io import BytesIO
from datetime import datetime

from .models import Event, Guest, Invitation, RSVP, CheckInLog


@login_required
def export_guest_list_csv(request, event_id):
    """Export guest list to CSV"""
    event = get_object_or_404(Event, id=event_id)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="guest_list_{event.name}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Email', 'Phone', 'Rank', 'Institution', 'RSVP Status', 'Plus Ones', 'Table', 'Seat', 'Checked In'])
    
    for invitation in event.invitations.select_related('guest').all():
        rsvp = getattr(invitation, 'rsvp', None)
        rsvp_status = rsvp.response if rsvp else 'No Response'
        plus_ones = rsvp.plus_ones if rsvp else 0
        
        writer.writerow([
            invitation.guest.full_name,
            invitation.guest.email,
            invitation.guest.phone or '',
            invitation.guest.rank or '',
            invitation.guest.institution or '',
            rsvp_status,
            plus_ones,
            invitation.table_number or '',
            invitation.seat_number or '',
            'Yes' if invitation.checked_in else 'No'
        ])
    
    return response


@login_required
def export_guest_list_excel(request, event_id):
    """Export guest list to Excel with formatting"""
    event = get_object_or_404(Event, id=event_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guest List"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    headers = ['Name', 'Email', 'Phone', 'Rank', 'Institution', 'RSVP Status', 'Plus Ones', 'Table', 'Seat', 'Checked In']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, invitation in enumerate(event.invitations.select_related('guest').all(), 2):
        rsvp = getattr(invitation, 'rsvp', None)
        rsvp_status = rsvp.response if rsvp else 'No Response'
        plus_ones = rsvp.plus_ones if rsvp else 0
        
        data = [
            invitation.guest.full_name,
            invitation.guest.email,
            invitation.guest.phone or '',
            invitation.guest.rank or '',
            invitation.guest.institution or '',
            rsvp_status,
            plus_ones,
            invitation.table_number or '',
            invitation.seat_number or '',
            'Yes' if invitation.checked_in else 'No'
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            
            # Color code RSVP status
            if col_num == 6:  # RSVP Status column
                if value == 'yes':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == 'no':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif value == 'maybe':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            # Color code check-in status
            if col_num == 10:  # Checked In column
                if value == 'Yes':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, color="006100")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="guest_list_{event.name}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def export_checkin_log_csv(request, event_id):
    """Export check-in log to CSV"""
    event = get_object_or_404(Event, id=event_id)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="checkin_log_{event.name}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Guest Name', 'Email', 'Check-In Time', 'Table', 'Seat'])
    
    for invitation in event.invitations.filter(checked_in=True).select_related('guest').order_by('check_in_time'):
        writer.writerow([
            invitation.guest.full_name,
            invitation.guest.email,
            invitation.check_in_time.strftime('%Y-%m-%d %H:%M:%S') if invitation.check_in_time else '',
            invitation.table_number or '',
            invitation.seat_number or ''
        ])
    
    return response


@login_required
def export_rsvp_report_csv(request, event_id):
    """Export RSVP report to CSV"""
    event = get_object_or_404(Event, id=event_id)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="rsvp_report_{event.name}_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Guest Name', 'Email', 'RSVP Status', 'Plus Ones', 'Dietary Restrictions', 'Notes', 'Responded At'])
    
    for invitation in event.invitations.select_related('guest').all():
        rsvp = getattr(invitation, 'rsvp', None)
        if rsvp:
            writer.writerow([
                invitation.guest.full_name,
                invitation.guest.email,
                rsvp.response,
                rsvp.plus_ones,
                rsvp.dietary_restrictions or '',
                rsvp.notes or '',
                rsvp.responded_at.strftime('%Y-%m-%d %H:%M:%S') if rsvp.responded_at else ''
            ])
        else:
            writer.writerow([
                invitation.guest.full_name,
                invitation.guest.email,
                'No Response',
                0,
                '',
                '',
                ''
            ])
    
    return response


@login_required
def export_seating_chart_pdf(request, event_id):
    """Export seating chart to PDF"""
    event = get_object_or_404(Event, id=event_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="seating_chart_{event.name}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#0066CC'),
        spaceAfter=30,
        alignment=1  # Center
    )
    
    title = Paragraph(f"Seating Chart - {event.name}", title_style)
    elements.append(title)
    
    subtitle = Paragraph(f"{event.date.strftime('%B %d, %Y at %I:%M %p')} - {event.location}", styles['Normal'])
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3*inch))
    
    # Get all tables
    tables = {}
    for invitation in event.invitations.filter(table_number__isnull=False).select_related('guest'):
        table_num = invitation.table_number
        if table_num not in tables:
            tables[table_num] = []
        tables[table_num].append(invitation)
    
    # Create table for each table number
    for table_num in sorted(tables.keys(), key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else 0, x)):
        # Table header
        table_title = Paragraph(f"<b>Table {table_num}</b>", styles['Heading2'])
        elements.append(table_title)
        elements.append(Spacer(1, 0.1*inch))
        
        # Guest list for this table
        data = [['Seat', 'Guest Name', 'Rank', 'RSVP', 'Checked In']]
        
        for invitation in sorted(tables[table_num], key=lambda x: x.seat_number or ''):
            rsvp = getattr(invitation, 'rsvp', None)
            rsvp_status = rsvp.response if rsvp else 'No Response'
            
            data.append([
                invitation.seat_number or '',
                invitation.guest.full_name,
                invitation.guest.rank or '',
                rsvp_status.upper(),
                '✓' if invitation.checked_in else ''
            ])
        
        # Create and style table
        t = Table(data, colWidths=[0.8*inch, 3*inch, 1.5*inch, 1.2*inch, 1*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066CC')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(t)
        elements.append(Spacer(1, 0.3*inch))
    
    # Build PDF
    doc.build(elements)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@login_required
def print_invitation_card(request, invitation_id):
    """Print invitation card with QR code and barcode"""
    from django.core.exceptions import PermissionDenied
    from reportlab.platypus import PageBreak
    
    invitation = get_object_or_404(Invitation, id=invitation_id)
    event = invitation.event
    
    # Permission check
    if not (request.user.is_staff or request.user.is_superuser or event.created_by == request.user):
        raise PermissionDenied
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invitation_card_{invitation.guest.full_name}_{event.name}.pdf"'
    
    buffer = BytesIO()
    # Use smaller page size for invitation card (6x4 inches)
    page_width = 6*inch
    page_height = 4*inch
    doc = SimpleDocTemplate(buffer, pagesize=(page_width, page_height), 
                           topMargin=0.3*inch, bottomMargin=0.3*inch,
                           leftMargin=0.3*inch, rightMargin=0.3*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'InvitationTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0066CC'),
        spaceAfter=10,
        alignment=1  # Center
    )
    
    name_style = ParagraphStyle(
        'GuestName',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.black,
        spaceAfter=8,
        alignment=1
    )
    
    detail_style = ParagraphStyle(
        'Details',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        alignment=1
    )
    
    # Title
    elements.append(Paragraph("You're Invited!", title_style))
    
    # Guest name with rank
    guest_name = f"{invitation.guest.rank} {invitation.guest.full_name}" if invitation.guest.rank else invitation.guest.full_name
    elements.append(Paragraph(f"<b>{guest_name}</b>", name_style))
    
    # Event details
    elements.append(Paragraph(f"<b>{event.name}</b>", detail_style))
    elements.append(Paragraph(f"📅 {event.date.strftime('%B %d, %Y at %I:%M %p')}", detail_style))
    elements.append(Paragraph(f"📍 {event.location}", detail_style))
    
    # Seating info if available
    if invitation.table_number:
        seating_info = f"Table {invitation.table_number}"
        if invitation.seat_number:
            seating_info += f", Seat {invitation.seat_number}"
        elements.append(Paragraph(f"🪑 {seating_info}", detail_style))
    
    elements.append(Spacer(1, 0.15*inch))
    
    # QR Code and Barcode section
    codes_data = []
    
    # Add QR code if available
    if invitation.qr_code and hasattr(invitation.qr_code, 'path') and os.path.exists(invitation.qr_code.path):
        try:
            qr_img = ReportLabImage(invitation.qr_code.path, width=1.2*inch, height=1.2*inch)
            qr_label = Paragraph("<b>Scan to RSVP</b>", ParagraphStyle('QRLabel', parent=styles['Normal'], fontSize=8, alignment=1))
            qr_cell = [[qr_img], [qr_label]]
            codes_data.append(Table(qr_cell, colWidths=[1.2*inch]))
        except:
            pass
    
    # Add barcode if available
    if invitation.barcode_image and hasattr(invitation.barcode_image, 'path') and os.path.exists(invitation.barcode_image.path):
        try:
            barcode_img = ReportLabImage(invitation.barcode_image.path, width=2*inch, height=0.8*inch)
            barcode_label = Paragraph("<b>Check-in Code</b>", ParagraphStyle('BarcodeLabel', parent=styles['Normal'], fontSize=8, alignment=1))
            barcode_cell = [[barcode_img], [barcode_label]]
            codes_data.append(Table(barcode_cell, colWidths=[2*inch]))
        except:
            pass
    
    # Create table with codes side by side if both available
    if codes_data:
        codes_table = Table([codes_data], colWidths=[2.5*inch] * len(codes_data))
        codes_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(codes_table)
    
    elements.append(Spacer(1, 0.1*inch))
    
    # Footer
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey)
    if event.rsvp_deadline:
        elements.append(Paragraph(f"Please RSVP by {event.rsvp_deadline.strftime('%B %d, %Y')}", footer_style))
    
    # Border around entire card
    def add_border(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#0066CC'))
        canvas.setLineWidth(2)
        canvas.rect(0.2*inch, 0.2*inch, page_width - 0.4*inch, page_height - 0.4*inch)
        canvas.restoreState()
    
    # Build PDF
    doc.build(elements, onFirstPage=add_border, onLaterPages=add_border)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@login_required
def print_event_invitation_cards(request, event_id):
    """Print invitation cards for all guests in an event"""
    from django.core.exceptions import PermissionDenied
    from reportlab.platypus import PageBreak
    
    event = get_object_or_404(Event, id=event_id)
    
    # Permission check
    if not (request.user.is_staff or request.user.is_superuser or event.created_by == request.user):
        raise PermissionDenied
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invitation_cards_{event.name}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    buffer = BytesIO()
    # Use card size
    page_width = 6*inch
    page_height = 4*inch
    doc = SimpleDocTemplate(buffer, pagesize=(page_width, page_height),
                           topMargin=0.3*inch, bottomMargin=0.3*inch,
                           leftMargin=0.3*inch, rightMargin=0.3*inch)
    
    all_elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'InvitationTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0066CC'),
        spaceAfter=10,
        alignment=1
    )
    
    name_style = ParagraphStyle(
        'GuestName',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.black,
        spaceAfter=8,
        alignment=1
    )
    
    detail_style = ParagraphStyle(
        'Details',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=6,
        alignment=1
    )
    
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=1, textColor=colors.grey)
    
    invitations = event.invitations.select_related('guest').order_by('guest__last_name', 'guest__first_name')
    
    for idx, invitation in enumerate(invitations):
        if idx > 0:
            all_elements.append(PageBreak())
        
        # Title
        all_elements.append(Paragraph("You're Invited!", title_style))
        
        # Guest name with rank
        guest_name = f"{invitation.guest.rank} {invitation.guest.full_name}" if invitation.guest.rank else invitation.guest.full_name
        all_elements.append(Paragraph(f"<b>{guest_name}</b>", name_style))
        
        # Event details
        all_elements.append(Paragraph(f"<b>{event.name}</b>", detail_style))
        all_elements.append(Paragraph(f"📅 {event.date.strftime('%B %d, %Y at %I:%M %p')}", detail_style))
        all_elements.append(Paragraph(f"📍 {event.location}", detail_style))
        
        # Seating info
        if invitation.table_number:
            seating_info = f"Table {invitation.table_number}"
            if invitation.seat_number:
                seating_info += f", Seat {invitation.seat_number}"
            all_elements.append(Paragraph(f"🪑 {seating_info}", detail_style))
        
        all_elements.append(Spacer(1, 0.15*inch))
        
        # QR Code and Barcode
        codes_data = []
        
        if invitation.qr_code and hasattr(invitation.qr_code, 'path') and os.path.exists(invitation.qr_code.path):
            try:
                qr_img = ReportLabImage(invitation.qr_code.path, width=1.2*inch, height=1.2*inch)
                qr_label = Paragraph("<b>Scan to RSVP</b>", ParagraphStyle('QRLabel', parent=styles['Normal'], fontSize=8, alignment=1))
                qr_cell = [[qr_img], [qr_label]]
                codes_data.append(Table(qr_cell, colWidths=[1.2*inch]))
            except:
                pass
        
        if invitation.barcode_image and hasattr(invitation.barcode_image, 'path') and os.path.exists(invitation.barcode_image.path):
            try:
                barcode_img = ReportLabImage(invitation.barcode_image.path, width=2*inch, height=0.8*inch)
                barcode_label = Paragraph("<b>Check-in Code</b>", ParagraphStyle('BarcodeLabel', parent=styles['Normal'], fontSize=8, alignment=1))
                barcode_cell = [[barcode_img], [barcode_label]]
                codes_data.append(Table(barcode_cell, colWidths=[2*inch]))
            except:
                pass
        
        if codes_data:
            codes_table = Table([codes_data], colWidths=[2.5*inch] * len(codes_data))
            codes_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            all_elements.append(codes_table)
        
        all_elements.append(Spacer(1, 0.1*inch))
        
        # Footer
        if event.rsvp_deadline:
            all_elements.append(Paragraph(f"Please RSVP by {event.rsvp_deadline.strftime('%B %d, %Y')}", footer_style))
    
    # Border function
    def add_border(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor('#0066CC'))
        canvas.setLineWidth(2)
        canvas.rect(0.2*inch, 0.2*inch, page_width - 0.4*inch, page_height - 0.4*inch)
        canvas.restoreState()
    
    # Build PDF
    doc.build(all_elements, onFirstPage=add_border, onLaterPages=add_border)
    
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response
